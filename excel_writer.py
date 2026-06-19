import hashlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CATEGORY_ORDER, CATEGORY_COLORS


def create_pdf_jump_html(record, link_dir):
    link_dir.mkdir(parents=True, exist_ok=True)

    unique_text = f"{record['source_path']}_{record['page_no']}_{record['category_index']}"
    file_hash = hashlib.md5(unique_text.encode("utf-8")).hexdigest()

    html_path = link_dir / f"link_{file_hash}.html"

    pdf_uri = Path(record["source_path"]).resolve().as_uri()
    page_no = record["page_no"]

    html = f"""
<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<title>PDF Link</title>
<script>
window.location.href = "{pdf_uri}#page={page_no}";
</script>
</head>
<body>
<p>PDFを開いています...</p>
<p><a href="{pdf_uri}#page={page_no}">開かない場合はこちらをクリック</a></p>
</body>
</html>
"""

    html_path.write_text(html, encoding="utf-8")
    return str(html_path.resolve())


def write_headers(ws, headers):
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)


def apply_style(ws):
    header_fill = PatternFill(
        start_color="D9EAF7",
        end_color="D9EAF7",
        fill_type="solid"
    )
    header_font = Font(bold=True)
    thin = Side(style="thin")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )
            cell.border = Border(
                left=thin,
                right=thin,
                top=thin,
                bottom=thin
            )

            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_num in range(1, ws.max_row + 1):
        ws.row_dimensions[row_num].height = 20


def color_by_category(ws, category_col):
    for row_num in range(2, ws.max_row + 1):
        category = ws.cell(row=row_num, column=category_col).value
        color = CATEGORY_COLORS.get(category)

        if not color:
            continue

        fill = PatternFill(
            start_color=color,
            end_color=color,
            fill_type="solid"
        )

        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_num, column=col).fill = fill


def create_comparison_sheet(wb, records):
    ws = wb.create_sheet("複数仕様書比較")

    source_files = sorted(set(record["source_file"] for record in records))

    headers = ["条件書項目"] + source_files
    write_headers(ws, headers)

    row = 2

    for category in CATEGORY_ORDER:
        ws.cell(row=row, column=1, value=category)

        for col, source_file in enumerate(source_files, start=2):
            pages = sorted({
                record["page_no"]
                for record in records
                if record["category"] == category
                and record["source_file"] == source_file
            })

            if pages:
                page_text = ", ".join([f"{p}P" for p in pages])
            else:
                page_text = "-"

            ws.cell(row=row, column=col, value=page_text)

        row += 1

    apply_style(ws)

    ws.column_dimensions["A"].width = 32

    for col in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 26

    color_by_category(ws, 1)


def create_excel(records, output_excel):
    wb = Workbook()

    output_path = Path(output_excel).resolve()
    link_dir = output_path.parent / "pdf_links"

    # =========================
    # 該当ページ一覧
    # =========================

    ws = wb.active
    ws.title = "該当ページ一覧"

    headers = [
        "No",
        "抽出元資料",
        "条件書項目",
        "該当ページ",
        "一言要約",
        "表",
        "PDFリンク",
        "確認"
    ]

    write_headers(ws, headers)

    row = 2

    for i, record in enumerate(records, start=1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=record["source_file"])
        ws.cell(row=row, column=3, value=record["category"])
        ws.cell(row=row, column=4, value=record["page_no"])
        ws.cell(row=row, column=5, value=record.get("summary", "関連ページ"))
        ws.cell(row=row, column=6, value=record["has_table"])

        link_cell = ws.cell(row=row, column=7, value="PDFを開く")
        link_cell.hyperlink = create_pdf_jump_html(record, link_dir)
        link_cell.style = "Hyperlink"

        ws.cell(row=row, column=8, value=record["check"])

        row += 1

    apply_style(ws)
    color_by_category(ws, 3)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 10

    # =========================
    # 複数仕様書比較シート
    # =========================

    create_comparison_sheet(wb, records)

    # =========================
    # 項目別シート
    # =========================

    for category in CATEGORY_ORDER:
        sheet_name = category.replace(".", "_").replace(" ", "")[:25]
        ws_cat = wb.create_sheet(sheet_name)

        cat_headers = [
            "No",
            "抽出元資料",
            "該当ページ",
            "一言要約",
            "表",
            "PDFリンク",
            "確認"
        ]

        write_headers(ws_cat, cat_headers)

        cat_records = [
            r for r in records
            if r["category"] == category
        ]

        row = 2

        for i, record in enumerate(cat_records, start=1):
            ws_cat.cell(row=row, column=1, value=i)
            ws_cat.cell(row=row, column=2, value=record["source_file"])
            ws_cat.cell(row=row, column=3, value=record["page_no"])
            ws_cat.cell(row=row, column=4, value=record.get("summary", "関連ページ"))
            ws_cat.cell(row=row, column=5, value=record["has_table"])

            link_cell = ws_cat.cell(row=row, column=6, value="PDFを開く")
            link_cell.hyperlink = create_pdf_jump_html(record, link_dir)
            link_cell.style = "Hyperlink"

            ws_cat.cell(row=row, column=7, value=record["check"])

            row += 1

        apply_style(ws_cat)

        ws_cat.column_dimensions["A"].width = 5
        ws_cat.column_dimensions["B"].width = 28
        ws_cat.column_dimensions["C"].width = 10
        ws_cat.column_dimensions["D"].width = 32
        ws_cat.column_dimensions["E"].width = 6
        ws_cat.column_dimensions["F"].width = 12
        ws_cat.column_dimensions["G"].width = 10

    # =========================
    # シート順整理
    # =========================

    ordered_sheet_names = [
        "該当ページ一覧",
        "複数仕様書比較"
    ]

    for category in CATEGORY_ORDER:
        ordered_sheet_names.append(
            category.replace(".", "_").replace(" ", "")[:25]
        )

    wb._sheets = [
        wb[name]
        for name in ordered_sheet_names
        if name in wb.sheetnames
    ]

    wb.save(output_excel)